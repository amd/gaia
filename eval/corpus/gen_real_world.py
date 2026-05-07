#!/usr/bin/env python3
"""
Synthetic real-world corpus generator for the GAIA eval suite.

Generates 19 documents under eval/corpus/real_world/ matching the paths
referenced by eval/scenarios/real_world/*.yaml. Every document contains the
ground-truth facts the judge expects, framed in original prose so we don't
ship copyrighted source material.

Idempotent: re-running rebuilds the corpus from the data structures below.

Usage:
    python eval/corpus/gen_real_world.py
"""

from __future__ import annotations

import os
from pathlib import Path

# -- Repo-relative root for the synthetic corpus -----------------------------
REPO_ROOT = Path(__file__).resolve().parents[2]
CORPUS_ROOT = REPO_ROOT / "eval" / "corpus" / "real_world"


# ============================================================================
# TEXT DOCUMENTS (15 of 19)
# ============================================================================
TEXT_DOCS: dict[str, str] = {}


# -- Financial: Alphabet 10-K excerpt ----------------------------------------
TEXT_DOCS[
    "financial/sec_10k_alphabet_2024_excerpt.txt"
] = """ALPHABET INC.
ANNUAL REPORT ON FORM 10-K — EXCERPT
For the Fiscal Year Ended December 31, 2024
(Synthetic excerpt for evaluation purposes)

ITEM 1. BUSINESS — OVERVIEW

Alphabet Inc. is a holding company whose consolidated subsidiaries include
Google LLC and a group of other businesses collectively referred to as
"Other Bets." This excerpt summarizes selected financial results and
disclosures from the fiscal year 2024 annual report. It is intended for
evaluation harness use only and does not reproduce the original filing.

CONSOLIDATED RESULTS — FISCAL YEAR 2024

For the full year ended December 31, 2024, Alphabet reported total revenues
of $350,018 million ($350.0 billion), an increase of approximately 14% from
the prior year. Growth was driven by continued strength in Google Search &
other, YouTube ads, and Google Cloud. Operating income, operating margin,
net income, and diluted earnings per share each improved year over year, as
described in Item 7 of this report.

Geographic mix remained broadly consistent with FY2023, with the United
States accounting for slightly less than half of total revenues and
international markets contributing the remainder.

SEGMENT RESULTS

Alphabet reports three segments: Google Services, Google Cloud, and Other
Bets. The discussion below highlights Google Cloud's fourth-quarter
performance, which the company called out as a key driver of full-year
growth.

  Google Cloud — Fourth Quarter 2024
    Revenue:                 $12.0 billion
    Year-over-year growth:   30%

The Q4 2024 result represented an acceleration from prior quarters and
reflected continued adoption of Google Cloud Platform infrastructure
services and AI-related offerings introduced during the year.

Google Services and Other Bets segment results for FY2024 are presented in
the segment information note within Item 8; per-segment quarterly
breakdowns for those segments are not reproduced in this excerpt.

HUMAN CAPITAL

As of December 31, 2024, Alphabet had 183,323 full-time employees worldwide.
The company continues to invest in technical talent across research,
engineering, and product roles.

ITEM 7. MANAGEMENT'S DISCUSSION AND ANALYSIS

Total revenues of $350,018 million for the year ended December 31, 2024
grew approximately 14% compared with FY2023, primarily reflecting growth in
advertising revenues across Google Services and acceleration in Google
Cloud. Foreign exchange effects were a modest headwind during the year.

Note on scope: this excerpt covers full-year and selected segment results
only. It does not include market data such as the closing share price on
any given trading day; share-price history is published separately by
exchanges and is not part of the 10-K filing.

END OF EXCERPT
"""


# -- Financial: Fed Rate Decision Nov 2024 -----------------------------------
TEXT_DOCS[
    "financial/fed_rate_decision_nov2024.txt"
] = """FEDERAL RESERVE PRESS RELEASE — SYNTHETIC EXCERPT
For Release: November 7, 2024

The Federal Open Market Committee (FOMC) concluded its two-day meeting on
November 7, 2024 with a decision on the federal funds target rate. The
following summary paraphrases the press release for evaluation purposes.

POLICY DECISION

The Committee decided to set the target range for the federal funds rate
at 4-1/2 to 4-3/4 percent (4.50% to 4.75%). This represented a 25
basis-point reduction from the prior target range. In supporting the
decision, the Committee cited continued progress on inflation toward the
2 percent objective and a labor market that, while still solid, had
shown some signs of cooling relative to earlier in the year.

VOTE

The vote on the policy action was unanimous: 12 in favor and 0 against.
All voting members of the Committee supported the 25 basis-point
reduction and the post-meeting policy statement language.

OUTLOOK STATEMENT

The Committee will continue to monitor the implications of incoming
information for the economic outlook. In assessing the appropriate stance
of monetary policy, the Committee will continue to monitor a wide range
of indicators, including readings on labor market conditions, inflation
pressures and inflation expectations, and financial and international
developments.

NOTE ON SCOPE

This press release reports the policy decision and vote count only. The
Committee's quarterly economic projections — including projections for
gross domestic product (GDP) growth, the unemployment rate, and inflation
— are published separately in the Summary of Economic Projections (SEP).
The SEP and dot-plot are not part of this rate-decision release.

END OF RELEASE
"""


# -- Financial: Treasury FY2024 budget results -------------------------------
TEXT_DOCS[
    "financial/treasury_fy2024_budget_results.txt"
] = """U.S. DEPARTMENT OF THE TREASURY
FINAL MONTHLY TREASURY STATEMENT — FY2024 RESULTS SUMMARY
(Synthetic excerpt for evaluation purposes)

OVERVIEW

This summary describes the U.S. federal government's budget results for
fiscal year 2024 (FY2024), which ended September 30, 2024. The figures
below are drawn from the final Monthly Treasury Statement and present
totals on a unified-budget basis.

HEADLINE RESULTS — FY2024

  Federal budget deficit:        $1.833 trillion
  Deficit as a share of GDP:     6.4%

The FY2024 deficit reflected continued growth in mandatory outlays,
elevated interest costs on the public debt, and discretionary spending
increases relative to receipts.

FEDERAL DEBT — END OF FY2024

  Federal debt held by the public:   $28.2 trillion
  Public debt as a share of GDP:     98%

Debt held by the public rose during the fiscal year as new Treasury
issuance financed the deficit. The 98% public-debt-to-GDP ratio represents
a multi-decade high.

REVENUE AND OUTLAY HIGHLIGHTS

Total federal receipts for FY2024 grew modestly relative to FY2023, led by
individual income tax receipts and corporate income tax receipts. Total
federal outlays increased somewhat faster, with the largest line items
remaining Social Security, Medicare, defense, and net interest on the
public debt. Net interest costs rose meaningfully during the year as
higher market rates flowed through the Treasury's portfolio of marketable
debt.

NOTE ON SCOPE

This document reports FY2024 results only. Prior-year comparisons by
fiscal-year totals (for example, FY2023 deficit figures) are not included
in this summary; those figures are available in the FY2023 final Monthly
Treasury Statement and the historical tables maintained separately by
Treasury.

END OF EXCERPT
"""


# -- Legal: Apache License 2.0 -----------------------------------------------
TEXT_DOCS[
    "legal/apache_license_2.0.txt"
] = """APACHE LICENSE, VERSION 2.0 — SYNTHETIC PARAPHRASE
(Reformatted summary intended for evaluation harness use only)

This text paraphrases the Apache License 2.0 for evaluation purposes. It
preserves the structure and section numbering of the original license so
that questions about specific clauses can be answered accurately, but the
prose is original.

SECTION 1. DEFINITIONS

Defines "License," "Licensor," "Legal Entity," "You," "Source," "Object,"
"Work," "Derivative Works," "Contribution," and "Contributor."

SECTION 2. GRANT OF COPYRIGHT LICENSE

Each Contributor hereby grants You a perpetual, worldwide, non-exclusive,
no-charge, royalty-free, irrevocable copyright license to reproduce,
prepare Derivative Works of, publicly display, publicly perform,
sublicense, and distribute the Work and such Derivative Works in Source
or Object form. This includes use for any purpose, including commercial
use.

SECTION 3. GRANT OF PATENT LICENSE

Each Contributor hereby grants You a perpetual, worldwide, non-exclusive,
no-charge, royalty-free, irrevocable patent license to make, have made,
use, offer to sell, sell, import, and otherwise transfer the Work.

If You institute patent litigation against any entity (including a
cross-claim or counterclaim in a lawsuit) alleging that the Work or a
Contribution incorporated within the Work constitutes direct or
contributory patent infringement, then any patent licenses granted to You
under this License for that Work shall terminate as of the date such
litigation is filed.

SECTION 4. REDISTRIBUTION

You may reproduce and distribute copies of the Work or Derivative Works
in any medium, with or without modifications, in Source or Object form,
provided that You meet the conditions enumerated in Section 4 (preserving
notices, providing a copy of the License, indicating modified files, and
retaining the contents of any NOTICE file).

SECTION 5. SUBMISSION OF CONTRIBUTIONS

Unless You explicitly state otherwise, any Contribution intentionally
submitted for inclusion in the Work shall be under the terms and
conditions of this License.

SECTION 6. TRADEMARKS

This License does not grant permission to use the trade names, trademarks,
service marks, or product names of the Licensor, except as required for
describing the origin of the Work and reproducing the content of the
NOTICE file. Trademark rights are explicitly excluded from the rights
granted by this License.

SECTION 7. DISCLAIMER OF WARRANTY

Unless required by applicable law or agreed to in writing, the Work is
provided on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
KIND, either express or implied.

SECTION 8. LIMITATION OF LIABILITY

In no event and under no legal theory shall any Contributor be liable to
You for damages, including any direct, indirect, special, incidental, or
consequential damages of any character arising as a result of this License
or out of the use or inability to use the Work.

SECTION 9. ACCEPTING WARRANTY OR ADDITIONAL LIABILITY

While redistributing the Work, You may choose to offer, and charge a fee
for, acceptance of support, warranty, indemnity, or other liability
obligations consistent with this License.

NOTE ON METADATA

This License text describes legal terms only. License-registry metadata
such as SPDX identifiers is maintained by separate registries and is not
included in this document.

END OF LICENSE
"""


# -- Legal: GDPR Article 17 --------------------------------------------------
TEXT_DOCS[
    "legal/gdpr_article_17_right_to_erasure.txt"
] = """REGULATION (EU) 2016/679 — GDPR
ARTICLE 17 — RIGHT TO ERASURE ('RIGHT TO BE FORGOTTEN')
(Synthetic paraphrase for evaluation purposes)

This document paraphrases Article 17 of the EU General Data Protection
Regulation, commonly cited as the "Right to Erasure" or "Right to be
Forgotten." It preserves clause structure for accurate lookup.

ARTICLE 17(1) — GROUNDS FOR ERASURE

The data subject shall have the right to obtain from the controller the
erasure of personal data concerning him or her without undue delay, and
the controller shall have the obligation to erase personal data without
undue delay where one of the following grounds applies:

  (a) The personal data are no longer necessary in relation to the
      purposes for which they were collected or otherwise processed.

  (b) The data subject withdraws consent on which the processing is based,
      and there is no other legal ground for the processing.

  (c) The data subject objects to the processing pursuant to Article
      21(1) and there are no overriding legitimate grounds, or the data
      subject objects to direct-marketing processing under Article 21(2).

  (d) The personal data have been unlawfully processed.

  (e) The personal data have to be erased for compliance with a legal
      obligation in Union or Member-State law to which the controller is
      subject.

  (f) The personal data have been collected in relation to the offer of
      information-society services referred to in Article 8(1) (services
      offered to a child).

In total, six grounds for erasure (a through f) are enumerated in
Article 17(1).

ARTICLE 17(2) — DATA THAT WAS MADE PUBLIC

Where the controller has made the personal data public and is obliged
under paragraph 1 to erase the personal data, the controller, taking
account of available technology and the cost of implementation, shall
take reasonable steps, including technical measures, to inform other
controllers that are processing the data that the data subject has
requested the erasure by such controllers of any links to, or copy or
replication of, those personal data.

ARTICLE 17(3) — EXCEPTIONS

Paragraphs 1 and 2 shall not apply to the extent that processing is
necessary for, among other reasons: exercising the right of freedom of
expression and information; compliance with a legal obligation requiring
processing or carrying out a task in the public interest; reasons of
public interest in the area of public health; archiving in the public
interest, scientific or historical research, or statistical purposes; or
the establishment, exercise, or defense of legal claims.

NOTE ON SCOPE

This document covers Article 17 only. Administrative fines and other
penalties for infringements of GDPR provisions are addressed separately
in Article 83 (general conditions for imposing administrative fines) and
related provisions; financial penalty figures are not specified in
Article 17 itself.

END OF ARTICLE 17
"""


# -- Legal: GitHub ToS excerpt -----------------------------------------------
TEXT_DOCS[
    "legal/github_terms_of_service_excerpt.txt"
] = """GITHUB TERMS OF SERVICE — SYNTHETIC EXCERPT
(Paraphrased policy summary for evaluation purposes only)

This excerpt paraphrases representative provisions of the GitHub Terms of
Service for evaluation harness use. It is not the binding agreement and
does not reflect the precise wording of any specific version.

A. ACCOUNT REQUIREMENTS

To create a personal account on GitHub, you must be at least 13 years of
age. The minimum age to create a GitHub account is 13 years old. If you
reside in a jurisdiction that requires a higher age of digital consent,
you must meet that higher local minimum to use the service.

Account creation is reserved for human individuals. Accounts must be
created and used by a human; "machine accounts" or accounts created by
automated methods are not permitted under these terms. Bots cannot create
accounts: a human must register and remain accountable for any account
used to operate automated software, bots, or scripts on the service.

You are responsible for keeping your account credentials secure and for
all activity that occurs under your account.

B. ACCEPTABLE USE

You may use GitHub only in compliance with these terms and all applicable
laws. You agree not to engage in activity that violates the Acceptable Use
Policies, including unlawful, abusive, or disruptive behavior on the
platform.

C. CONTENT AND INTELLECTUAL PROPERTY

You retain ownership of content you create and upload. By uploading
content to GitHub, you grant the company a license sufficient to host,
display, and process that content for the purpose of operating the
service.

D. TERMINATION

GitHub may suspend or terminate accounts that violate these terms,
including accounts created by automated methods or that misrepresent the
identity of their human owner.

E. SCOPE OF THIS DOCUMENT

This terms-of-service text covers user obligations and permitted use
only. Pricing for paid plans (including GitHub Enterprise) is published
separately on the GitHub pricing page and is not part of the Terms of
Service document.

END OF EXCERPT
"""


# -- Legal: MIT License ------------------------------------------------------
TEXT_DOCS[
    "legal/mit_license.txt"
] = """MIT LICENSE — SYNTHETIC PARAPHRASE
(Reformatted summary for evaluation harness use only)

This document paraphrases the MIT License for evaluation purposes. The
substantive permissions and the single condition are preserved.

Copyright (c) <year> <copyright holder>

Permission is hereby granted, free of charge, to any person obtaining a
copy of this software and associated documentation files (the
"Software"), to deal in the Software without restriction. Permitted uses
include, without limitation, the rights to use, copy, modify, merge,
publish, distribute, sublicense, and/or sell copies of the Software, and
to permit persons to whom the Software is furnished to do so, including
for commercial purposes. Commercial use is allowed.

CONDITION

The MIT License imposes a single condition on the permissions granted
above: the above copyright notice and this permission notice shall be
included in all copies or substantial portions of the Software.

That single notice-retention requirement is the only condition the MIT
License places on use, copying, modification, and redistribution. No
additional conditions, share-alike obligations, or attribution
requirements beyond the notice are imposed.

WARRANTY DISCLAIMER

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS
OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT.
IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY
CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

NOTE ON PATENT GRANT

Unlike the Apache License 2.0, the MIT License does not include an
explicit patent grant. The MIT text grants only the copyright permissions
described above; any patent rights are not addressed by this license.
There is no explicit patent grant in the MIT License.

END OF LICENSE
"""


# -- Government: BLS employment situation Dec 2025 ---------------------------
TEXT_DOCS[
    "government/bls_employment_december_2025.txt"
] = """U.S. BUREAU OF LABOR STATISTICS
THE EMPLOYMENT SITUATION — DECEMBER 2025
(Synthetic summary for evaluation purposes)

The Bureau of Labor Statistics published the following summary of U.S.
labor-market conditions for December 2025. Figures are seasonally
adjusted and based on the Current Employment Statistics and Current
Population Survey programs.

HEADLINE FIGURES

  Total nonfarm payroll employment, change in December 2025:  +50,000
  Unemployment rate, December 2025:                            4.4%

Nonfarm payroll employment rose by 50,000 in December 2025. The
unemployment rate was 4.4%, little changed from the prior month.

INDUSTRY DETAIL

Job gains in December 2025 were concentrated in service-providing
industries. The leisure and hospitality sector added the most jobs of any
industry, gaining +47,000 in December. Health care and social assistance
followed with +38,500 jobs added. Professional and business services and
private education also contributed modestly to the headline gain.
Manufacturing employment was little changed, while retail trade
employment edged down.

WAGES

Average hourly earnings for all employees on private nonfarm payrolls
rose during the month. Year-over-year wage growth was 3.8%.

NOTE ON DEMOGRAPHIC DETAIL

This summary covers aggregate national-level employment, unemployment,
and earnings indicators. Detailed demographic breakdowns by race or
ethnicity (for example, the unemployment rate for Black or African
American workers) are published in the accompanying detailed tables and
the BLS demographic data products. They are not included in this
employment-situation summary text.

END OF SUMMARY
"""


# -- Government: NIST CSF 2.0 overview ---------------------------------------
TEXT_DOCS[
    "government/nist_csf_2_overview.txt"
] = """NIST CYBERSECURITY FRAMEWORK 2.0 — FRAMEWORK OVERVIEW
(Synthetic summary for evaluation purposes)

The NIST Cybersecurity Framework (CSF) 2.0 is a voluntary framework
published by the U.S. National Institute of Standards and Technology to
help organizations manage and reduce cybersecurity risk. Version 2.0
expands and refines the original framework first published in 2014 and
updated as version 1.1 in 2018.

RELEASE

NIST CSF 2.0 was released on February 26, 2024. The 2.0 release marks the
first major revision since version 1.1.

WHAT'S NEW IN 2.0

The most significant structural change in CSF 2.0 is the addition of a
new core function. Version 1.1 included five core functions; version 2.0
introduces a sixth core function, GOVERN (abbreviated GV), which was not
present in version 1.1. The new GOVERN function elevates organizational
governance, risk-management strategy, and supply-chain considerations to
the top level of the framework.

CORE FUNCTIONS (CSF 2.0)

CSF 2.0 organizes cybersecurity activities into 6 core functions:
  1. Govern   (GV)  — new in 2.0
  2. Identify (ID)
  3. Protect  (PR)
  4. Detect   (DE)
  5. Respond  (RS)
  6. Recover  (RC)

CATEGORIES AND SUBCATEGORIES

The 6 core functions are decomposed into categories and subcategories.
CSF 2.0 contains a total of 106 subcategories across the six functions,
each describing a specific cybersecurity outcome that organizations may
work toward.

INTENDED AUDIENCE

CSF 2.0 is intended for use by organizations of all sizes and sectors,
including critical infrastructure operators, small businesses, and
non-U.S. organizations. The framework is technology-neutral and is
designed to be used alongside sector-specific guidance and other risk
management resources.

NOTE ON PROCESS HISTORY

This overview summarizes the framework structure and release. Process
artifacts associated with the development of CSF 2.0 — for example, the
length of the public comment period preceding final publication — are
documented in the separate development history materials maintained by
NIST and are not included in this overview.

END OF OVERVIEW
"""


# -- Scientific: Attention is All You Need -----------------------------------
TEXT_DOCS[
    "scientific/attention_transformer_arxiv.txt"
] = """ATTENTION IS ALL YOU NEED — PAPER SUMMARY
(Synthetic paraphrase for evaluation purposes)

This document summarizes the research paper commonly referred to as
"Attention Is All You Need." It is paraphrased for evaluation harness use
and is not a reproduction of the published article.

AUTHORS

The paper was written by Ashish Vaswani, Noam Shazeer, Niki Parmar,
Jakob Uszkoreit, Llion Jones, Aidan N. Gomez, Lukasz Kaiser, and Illia
Polosukhin. The first two authors are Ashish Vaswani and Noam Shazeer.

ARCHITECTURE

The paper proposes a sequence-transduction model called the Transformer.
The Transformer is based solely on attention mechanisms; it dispenses
entirely with recurrence and convolutions. The architecture stacks
self-attention and position-wise feed-forward layers in an encoder-decoder
configuration, with multi-head attention used both within the encoder and
decoder and as the cross-attention bridge between them. Because the model
contains no recurrent or convolutional units, sinusoidal positional
encodings are added to the input embeddings to inject sequence-order
information.

KEY RESULTS — WMT 2014 TRANSLATION BENCHMARKS

The Transformer was evaluated on the WMT 2014 English-to-German and
English-to-French machine-translation tasks. The reported results are:

  WMT 2014 English-to-German: 28.4 BLEU
  WMT 2014 English-to-French: 41.8 BLEU

The 41.8 BLEU score on English-to-French was the new single-model
state-of-the-art at publication, surpassing previous best results from
recurrent and convolutional sequence-to-sequence systems while requiring
substantially less training time.

TRAINING DETAILS

The Transformer was trained on standard WMT 2014 datasets using the Adam
optimizer with a custom learning-rate schedule (warmup followed by
inverse-square-root decay). Training was performed on 8 GPUs over the
course of several days. Regularization included residual dropout, label
smoothing, and attention dropout.

NOTE ON SCOPE

The benchmark results in this paper cover English-to-German and
English-to-French only. Results for other language pairs (for example,
English-to-Czech translation on WMT 2014) are not reported in this paper.

END OF SUMMARY
"""


# -- Scientific: CDC flu 2023-2024 -------------------------------------------
TEXT_DOCS[
    "scientific/cdc_influenza_2023_2024.txt"
] = """CDC SEASONAL INFLUENZA SURVEILLANCE — 2023-2024 SEASON SUMMARY
(Synthetic paraphrase for evaluation purposes)

The U.S. Centers for Disease Control and Prevention (CDC) tracks
influenza activity year-round through ILINet, FluView, and laboratory
surveillance networks. This document summarizes selected end-of-season
findings for the 2023-2024 influenza season.

SEVERITY CLASSIFICATION

CDC classified the 2023-2024 influenza season as moderately severe across
all age groups. The classification reflects the combination of outpatient
illness intensity, hospitalization rates, and mortality observed during
the season.

OUTPATIENT ILLNESS

Outpatient visits for influenza-like illness (ILI) rose sharply in late
2023, with the percentage of laboratory specimens testing positive for
influenza peaking during the week ending December 30, 2023. The peak
outpatient ILI positivity rate that week reached 18.3%.

PEDIATRIC MORTALITY

A total of 197 laboratory-confirmed influenza-associated pediatric deaths
were reported during the 2023-2024 season. This represented the
second-highest pediatric flu death count since CDC began collecting
pediatric flu mortality data in 2004; only one prior season exceeded this
total.

VIRUS COMPOSITION

Influenza A(H1N1)pdm09 viruses predominated early and through much of the
season, with influenza A(H3N2) and influenza B/Victoria-lineage activity
also detected. Antiviral susceptibility remained high throughout the
season.

NOTE ON SCOPE

This summary reports surveillance metrics on illness, hospitalization,
and mortality. Vaccination coverage rates by age group — including
estimates of the percentage of U.S. adults who received an influenza
vaccine during the 2023-2024 season — are published separately in CDC's
FluVaxView products and are not included in this surveillance summary.

END OF SUMMARY
"""


# -- Technical: Python 3.11 What's New ---------------------------------------
TEXT_DOCS[
    "technical/python311_whats_new.txt"
] = """WHAT'S NEW IN PYTHON 3.11 — SYNTHETIC SUMMARY
(Paraphrased changelog for evaluation purposes)

Python 3.11 was released on October 24, 2022. The release focused on
performance, error reporting, and standard-library additions. Below is a
condensed list of representative changes; the full changelog is published
separately on python.org.

PERFORMANCE

CPython 3.11 is approximately 1.25x faster (about 25% faster) than
CPython 3.10 across a broad range of workloads, as measured with the
pyperformance benchmark suite. The performance gains come from work in
the "Faster CPython" project, including specializing adaptive
interpreter, smaller and cheaper Python frames, and a "zero-cost"
exception model.

EXCEPTION GROUPS AND except*

PEP 654 introduces ExceptionGroup and BaseExceptionGroup along with the
new "except*" syntax for handling multiple unrelated exceptions raised
together (for example, by concurrent code).

FINER-GRAINED ERROR LOCATIONS

Tracebacks now point to the specific expression on a line that caused an
error, instead of just the line number, making complex one-liners easier
to debug.

NEW STANDARD-LIBRARY MODULE — TOML SUPPORT

Python 3.11 added a new standard-library module for parsing TOML files:
tomllib. The module was introduced by PEP 680 and provides a read-only
TOML parser implemented in pure Python. tomllib makes parsing
pyproject.toml and other TOML configuration files possible without
third-party dependencies.

TYPING IMPROVEMENTS

Several typing improvements landed in 3.11, including PEP 646 (variadic
generics), PEP 655 (Required and NotRequired in TypedDict), PEP 673
(Self type), and PEP 675 (LiteralString).

ASYNCIO TASKGROUP

asyncio.TaskGroup is added to manage groups of concurrent tasks with
structured concurrency semantics that interoperate with the new
ExceptionGroup type.

NOTE ON SCOPE

This document covers Python 3.11 only. Information about subsequent
releases — including the release date of Python 3.12 — is not included
in this Python 3.11 changelog and is published separately when those
versions are released.

END OF DOCUMENT
"""


# -- Technical: Raspberry Pi 4 datasheet -------------------------------------
TEXT_DOCS[
    "technical/raspberry_pi4_specifications.txt"
] = """RASPBERRY PI 4 MODEL B — PRODUCT SPECIFICATIONS
(Synthetic datasheet excerpt for evaluation purposes)

The Raspberry Pi 4 Model B is a single-board computer produced by the
Raspberry Pi Foundation. The specifications below summarize key technical
characteristics of the board.

SYSTEM-ON-CHIP (SoC)

The Raspberry Pi 4 uses a Broadcom BCM2711 system-on-chip. The SoC
contains a quad-core ARM Cortex-A72 (ARM v8) 64-bit application processor
clocked at 1.5 GHz, paired with a VideoCore VI GPU and integrated memory
and I/O controllers.

MEMORY

The Raspberry Pi 4 ships in four RAM variants:
  1 GB LPDDR4-3200 SDRAM
  2 GB LPDDR4-3200 SDRAM
  4 GB LPDDR4-3200 SDRAM
  8 GB LPDDR4-3200 SDRAM

All four variants use LPDDR4-3200 SDRAM. RAM is soldered to the board and
is not user-upgradeable.

NETWORKING

  Gigabit Ethernet (1 Gbps, RJ45)
  Dual-band 802.11ac wireless LAN (2.4 GHz and 5 GHz)
  Bluetooth 5.0, BLE

USB AND I/O

  2x USB 3.0 ports
  2x USB 2.0 ports
  2x micro-HDMI ports (each up to 4Kp60)
  1x USB-C port (used for power input only)
  40-pin GPIO header (compatible with earlier Raspberry Pi boards)

POWER

The board is powered via the USB-C port at 5 V / 3 A (15 W). A
Raspberry Pi-branded USB-C power supply is recommended.

OPERATING SYSTEM

The board is supported by Raspberry Pi OS (Debian-based) and a number of
third-party Linux distributions, as well as several Unix-like and
embedded operating systems.

NOTE ON COMMERCIAL INFORMATION

This datasheet describes hardware specifications only. Retail pricing for
the Raspberry Pi 4 (including the 8 GB model) is set by authorized
resellers and is not part of this technical specification document.

END OF DATASHEET
"""


# -- Technical: RFC 7231 -----------------------------------------------------
TEXT_DOCS[
    "technical/rfc7231_http_semantics.txt"
] = """RFC 7231 — HYPERTEXT TRANSFER PROTOCOL (HTTP/1.1): SEMANTICS AND CONTENT
(Synthetic paraphrase for evaluation purposes)

This document paraphrases selected sections of RFC 7231, which defines
the semantics of HTTP/1.1 messages, including request methods, request
header fields, response status codes, and response header fields.

REQUEST METHODS — REQUIRED METHODS

A general-purpose HTTP/1.1 server is required to support the GET and HEAD
methods. All other request methods (such as POST, PUT, DELETE, OPTIONS,
TRACE, and CONNECT) are optional and may be supported when appropriate
for the resource. GET and HEAD are the only required methods.

REQUEST METHODS — SAFE METHODS

A request method is considered "safe" when its semantics are read-only:
the client expects the server not to take any state-changing action as
a direct effect of the request. RFC 7231 classifies the following methods
as safe:

  GET
  HEAD
  OPTIONS
  TRACE

Other methods, such as POST, PUT, DELETE, and CONNECT, are not safe
because they may have side effects on the resource or on the server.

IDEMPOTENT METHODS

A request method is considered "idempotent" if multiple identical
requests have the same intended effect on the server as a single such
request. PUT, DELETE, and the safe methods (GET, HEAD, OPTIONS, TRACE)
are idempotent. POST is not, in general, idempotent.

CONTENT NEGOTIATION

RFC 7231 defines the proactive (server-driven) and reactive (agent-driven)
content negotiation models, including the Accept, Accept-Charset,
Accept-Encoding, and Accept-Language request header fields.

STATUS CODES

The document defines the meaning of the standard HTTP status code
classes (1xx informational, 2xx successful, 3xx redirection, 4xx client
error, 5xx server error) and the precise semantics of common codes such
as 200 OK, 201 Created, 204 No Content, 301 Moved Permanently, 304 Not
Modified, 400 Bad Request, 404 Not Found, and 500 Internal Server Error.

NOTE ON REQUEST BODY LIMITS

RFC 7231 does not define a maximum request body size. The maximum size
of a request payload is implementation-specific and may be configured by
individual servers, gateways, or intermediaries. Any specific byte limit
on request bodies is therefore outside the scope of this specification.

END OF EXCERPT
"""


# -- Technical: USB 2.0 spec overview ----------------------------------------
TEXT_DOCS[
    "technical/usb20_specification_overview.txt"
] = """UNIVERSAL SERIAL BUS REVISION 2.0 — SPECIFICATION OVERVIEW
(Synthetic excerpt for evaluation purposes)

This overview paraphrases selected sections of the USB 2.0 specification.
It is intended for evaluation harness use and is not the binding
specification document.

DATA TRANSFER MODES

USB 2.0 defines three data transfer modes, distinguished by their maximum
signaling rates:

  Low Speed:    1.5 Mbit/s
  Full Speed:   12 Mbit/s
  High Speed:   480 Mbit/s

The High Speed mode at 480 Mbit/s was the headline addition introduced
by USB 2.0 over USB 1.1.

BUS TOPOLOGY

USB uses a tiered-star topology rooted at a single host controller. A
USB 2.0 host controller can address up to 127 devices simultaneously
(devices are assigned 7-bit addresses; one address is reserved). Hubs are
permitted to provide additional downstream ports, up to a maximum of five
non-root tiers between the host and any device.

CABLES — HIGH SPEED

The maximum cable length for a USB 2.0 High Speed segment between a hub
and a device, or between two hubs, is 5 meters. This 5-meter limit is
the only cable-length value explicitly stated in this overview.

POWER

USB 2.0 specifies a bus voltage of 5 V (+/- 5%). A standard downstream
port supplies up to 100 mA at enumeration and up to 500 mA after
configuration. Self-powered hubs may supply 500 mA per downstream port.

CONNECTORS

USB 2.0 defines Standard-A and Standard-B receptacles and plugs, along
with the Mini-B variant for compact devices. The Micro-B and Type-C
connectors common on later devices are defined in subsequent USB
specifications.

NOTE ON SCOPE

This overview lists the 5-meter maximum cable length for High Speed mode
only. Maximum cable lengths for the Low Speed and Full Speed modes are
not separately stated in this overview.

END OF OVERVIEW
"""


# ============================================================================
# XLSX DOCUMENTS (4 of 19)
# ============================================================================
# We build them with openpyxl. The RAG pipeline in
# src/gaia/rag/sdk.py::_extract_text_from_xlsx flattens cells per-row using the
# header row as column names. So we put facts into clearly-labeled rows.


def _ws_set(ws, rows: list[list]) -> None:
    """Write a list of rows into the active worksheet."""
    for row in rows:
        ws.append(row)


def build_company_financials_xlsx(out_path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    # Default sheet renamed
    income = wb.active
    income.title = "Income Statement"
    # Net-income figures are after-tax bottom-line numbers (consistent with
    # FY2024 total $29,461,000 the judge expects). Operating income is the
    # pre-tax-and-non-operating line above.
    _ws_set(
        income,
        [
            ["Meridian Technology Solutions — FY2024 Income Statement"],
            ["Currency: USD; Period: fiscal year ended December 31, 2024"],
            [],
            ["Period", "Revenue", "Cost of Revenue", "Gross Profit", "Operating Expenses", "Operating Income", "Net Income"],
            ["Q1 2024", 34_120_000, 13_648_000, 20_472_000, 11_500_000, 8_972_000, 6_120_000],
            ["Q2 2024", 36_540_000, 14_616_000, 21_924_000, 11_900_000, 10_024_000, 6_840_000],
            ["Q3 2024", 38_250_000, 15_300_000, 22_950_000, 12_200_000, 10_750_000, 7_165_000],
            ["Q4 2024", 47_320_000, 18_928_000, 28_392_000, 13_800_000, 14_592_000, 9_336_000],
            ["FY2024 Total", 156_230_000, 62_492_000, 93_738_000, 49_400_000, 44_338_000, 29_461_000],
            [],
            ["Note: FY2024 total revenue is $156,230,000."],
            ["Note: Q4 2024 revenue is $47,320,000."],
            ["Note: FY2024 net income is $29,461,000 (after non-operating items and taxes)."],
            [],
            ["No forward guidance or 2025 revenue projections are included in this workbook."],
        ],
    )

    bs = wb.create_sheet("Balance Sheet")
    _ws_set(
        bs,
        [
            ["Meridian Technology Solutions — Balance Sheet"],
            ["As of December 31, 2024 (USD)"],
            [],
            ["Line Item", "Amount"],
            ["Cash and equivalents", 42_300_000],
            ["Accounts receivable", 28_500_000],
            ["Inventory", 12_400_000],
            ["Property, plant & equipment, net", 64_810_000],
            ["Goodwill and intangibles", 38_000_000],
            ["Total assets", 186_010_000],
            [],
            ["Accounts payable", 18_900_000],
            ["Accrued liabilities", 11_700_000],
            ["Long-term debt", 35_000_000],
            ["Total liabilities", 65_600_000],
            [],
            ["Total stockholders' equity", 120_410_000],
            ["Total liabilities and equity", 186_010_000],
            [],
            ["Note: Total assets as of December 31, 2024 are $186,010,000."],
            ["Note: This balance sheet does not include forward-looking 2025 projections."],
        ],
    )

    wb.save(out_path)


def build_department_budget_xlsx(out_path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    bva = wb.active
    bva.title = "Budget vs Actual"
    # Sales & Marketing: budget 12,900,000 actual 13,096,000 -> +196,000 +1.52%
    # Engineering:      budget 22,000,000 actual 22,260,000 -> +260,000 +1.18%
    # Operations:        budget 8,000,000  actual 7,994,000 -> -6,000  -0.075%
    # Customer Support:  budget 4,200,000  actual 4,242,000 -> +42,000 +1.00%
    # G&A:               budget 5,500,000  actual 5,553,000 -> +53,000 +0.96%
    # Human Resources:   budget 2,400,000  actual 2,422,000 -> +22,000 +0.92%
    _ws_set(
        bva,
        [
            ["Meridian Technology Solutions — FY2024 Budget vs Actual by Department"],
            ["Currency: USD; Variance % = (Actual - Budget) / Budget"],
            [],
            ["Department", "FY2024 Budget", "FY2024 Actual", "Variance ($)", "Variance (%)"],
            ["Engineering", 22_000_000, 22_260_000, 260_000, 0.0118],
            ["Sales & Marketing", 12_900_000, 13_096_000, 196_000, 0.0152],
            ["Operations", 8_000_000, 7_994_000, -6_000, -0.00075],
            ["Customer Support", 4_200_000, 4_242_000, 42_000, 0.0100],
            ["G&A", 5_500_000, 5_553_000, 53_000, 0.0096],
            ["Human Resources", 2_400_000, 2_422_000, 22_000, 0.0092],
            [],
            ["Note: Sales & Marketing is the most over-budget department at +1.52% (over by $196,000)."],
            ["Note: Engineering is over budget by +1.18% (over by $260,000) — second most over-budget by percent."],
            ["Note: Operations is the only department that came in under budget — under by $6,000 (-0.075%)."],
            ["No salary or per-employee compensation data is included in this workbook."],
        ],
    )

    hc = wb.create_sheet("Headcount")
    _ws_set(
        hc,
        [
            ["Meridian Technology Solutions — FY2024 Quarterly Headcount by Department"],
            [],
            ["Department", "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"],
            ["Engineering", 132, 136, 139, 142],
            ["Sales & Marketing", 70, 72, 74, 76],
            ["Operations", 48, 49, 50, 52],
            ["Customer Support", 36, 38, 40, 42],
            ["G&A", 22, 22, 23, 24],
            ["Human Resources", 14, 15, 16, 16],
            ["Total", 322, 332, 342, 352],
            [],
            ["Note: Q4 2024 total headcount is 352 employees."],
            ["Note: Engineering is the largest department in Q4 2024 with 142 employees."],
            ["Note: This sheet contains headcount counts only. No salary or compensation figures are present."],
        ],
    )

    summary = wb.create_sheet("Summary")
    _ws_set(
        summary,
        [
            ["FY2024 Budget Summary — Meridian Technology Solutions"],
            [],
            ["Metric", "Value"],
            ["Total FY2024 budget (all departments)", 55_000_000],
            ["Total FY2024 actual (all departments)", 55_567_000],
            ["Net variance (over)", 567_000],
            ["Departments over budget", 5],
            ["Departments under budget", 1],
            ["Most over-budget department (by %)", "Sales & Marketing"],
            ["Most over-budget department (variance %)", 0.0152],
            ["Only under-budget department", "Operations"],
            ["Largest department (Q4 2024)", "Engineering"],
            ["Largest department headcount (Q4 2024)", 142],
            ["Total Q4 2024 headcount", 352],
            [],
            ["Note: This workbook contains no salary or compensation data in any sheet."],
        ],
    )

    wb.save(out_path)


def build_product_inventory_xlsx(out_path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    inv = wb.active
    inv.title = "Inventory"

    # 30 products. SKU prefixes: COMP- for components/computing, ELEC- for
    # consumer electronics, ACCS- for accessories.
    # Required ground-truth rows:
    #   COMP-3001  AMD Ryzen 9 7950X CPU       stock 76
    #   COMP-3004  RX 7900 XTX GPU 24GB
    #   ACCS-5006  USB-C Docking Station       margin ≈ 48.0%  (highest)
    products = [
        # SKU,         Name,                              Category,    Stock, Reorder, Cost,    Price
        ("COMP-3001", "AMD Ryzen 9 7950X CPU",            "CPU",         76,     20,    420.00,  599.00),
        ("COMP-3002", "Intel Core i9-14900K CPU",         "CPU",         54,     20,    430.00,  589.00),
        ("COMP-3003", "RTX 4090 GPU 24GB",                "GPU",         18,     10,   1450.00, 1899.00),
        ("COMP-3004", "RX 7900 XTX GPU 24GB",             "GPU",         24,     10,    690.00,  949.00),
        ("COMP-3005", "RTX 4070 GPU 12GB",                "GPU",         42,     15,    420.00,  579.00),
        ("COMP-3006", "RX 7600 GPU 8GB",                  "GPU",         60,     20,    210.00,  289.00),
        ("COMP-3007", "32GB DDR5-6000 RAM Kit",           "Memory",     180,     50,     85.00,  139.00),
        ("COMP-3008", "16GB DDR5-5600 RAM Kit",           "Memory",     220,     60,     45.00,   79.00),
        ("COMP-3009", "1TB NVMe Gen4 SSD",                "Storage",    140,     40,     65.00,  109.00),
        ("COMP-3010", "2TB NVMe Gen4 SSD",                "Storage",     90,     30,    115.00,  189.00),
        ("ELEC-4001", "55-inch 4K OLED TV",               "Television",  22,     10,    900.00, 1399.00),
        ("ELEC-4002", "65-inch 4K OLED TV",               "Television",  16,      8,   1300.00, 1899.00),
        ("ELEC-4003", "75-inch 4K QLED TV",               "Television",  12,      6,   1100.00, 1599.00),
        ("ELEC-4004", "27-inch 4K Monitor",               "Monitor",     54,     20,    320.00,  479.00),
        ("ELEC-4005", "32-inch 4K Monitor",               "Monitor",     38,     15,    410.00,  599.00),
        ("ELEC-4006", "14-inch Ultrabook Laptop",         "Laptop",      40,     15,    780.00, 1099.00),
        ("ELEC-4007", "16-inch Workstation Laptop",       "Laptop",      28,     10,   1450.00, 1999.00),
        ("ELEC-4008", "13-inch Convertible Laptop",       "Laptop",      36,     12,    690.00,  949.00),
        ("ELEC-4009", "Flagship Smartphone 256GB",        "Phone",       55,     20,    620.00,  899.00),
        ("ELEC-4010", "Mid-range Smartphone 128GB",       "Phone",       70,     25,    280.00,  429.00),
        ("ELEC-4011", "Wireless Earbuds Pro",             "Phone",       95,     30,     90.00,  149.00),
        ("ACCS-5001", "Mechanical Keyboard 75%",          "Accessory",   80,     30,     58.00,   99.00),
        ("ACCS-5002", "Wireless Mouse Pro",               "Accessory",  120,     40,     32.00,   59.00),
        ("ACCS-5003", "USB-C Hub 7-in-1",                 "Accessory",   90,     30,     22.00,   39.00),
        ("ACCS-5004", "27-inch Monitor Stand",            "Accessory",   60,     20,     38.00,   69.00),
        ("ACCS-5005", "Webcam 1080p",                     "Accessory",   75,     25,     34.00,   59.00),
        # ACCS-5006 highest margin: cost 78, price 149.99 -> margin 47.997%
        ("ACCS-5006", "USB-C Docking Station",            "Accessory",   65,     25,     78.00,  149.99),
        ("ACCS-5007", "External SSD 1TB",                 "Accessory",   85,     30,     58.00,   99.00),
        ("ACCS-5008", "Noise-Cancelling Headphones",      "Accessory",   50,     20,    120.00,  199.00),
        ("ACCS-5009", "Standing Desk Mat",                "Accessory",   42,     15,     22.00,   39.00),
    ]
    assert len(products) == 30

    inv.append(["Meridian Electronics Distribution — Master Inventory"])
    inv.append([])
    inv.append(["SKU", "Product Name", "Category", "Stock", "Reorder Point", "Cost", "Price"])
    for p in products:
        inv.append(list(p))

    pl = wb.create_sheet("Price List")
    pl.append(["Meridian Electronics Distribution — Public Price List & Margin"])
    pl.append([])
    pl.append(["SKU", "Product Name", "Cost", "Price", "Margin %", "In Stock"])
    margins = []
    for sku, name, _cat, stock, _ro, cost, price in products:
        margin_pct = (price - cost) / price
        margins.append((sku, name, margin_pct))
        pl.append([sku, name, cost, price, round(margin_pct * 100, 2), stock])
    pl.append([])
    pl.append([
        "Note: Highest-margin product is ACCS-5006 USB-C Docking Station at approximately 48.0% margin."
    ])
    pl.append([
        "Note: All 30 SKUs currently have stock above their reorder point — no items are below reorder point."
    ])
    # Sanity-check: highest-margin row must indeed be ACCS-5006.
    margins_sorted = sorted(margins, key=lambda x: -x[2])
    assert margins_sorted[0][0] == "ACCS-5006", margins_sorted[0]

    lk = wb.create_sheet("Lookup")
    lk.append(["Quick-reference lookup for inventory operations"])
    lk.append([])
    lk.append(["Question", "Answer"])
    lk.append(["RX 7900 XTX GPU 24GB SKU", "COMP-3004"])
    lk.append(["AMD Ryzen 9 7950X CPU SKU", "COMP-3001"])
    lk.append(["AMD Ryzen 9 7950X CPU stock", 76])
    lk.append(["Highest-margin SKU", "ACCS-5006"])
    lk.append(["Highest-margin product", "USB-C Docking Station"])
    lk.append(["Highest-margin percentage (approx)", "48.0%"])
    lk.append(["Items currently below reorder point", 0])
    lk.append(["Total distinct SKUs in inventory", 30])
    lk.append([])
    lk.append([
        "Note: This workbook does not include warranty information for any product."
    ])
    wb.save(out_path)

    # Cross-check that the asserted stock value for COMP-3001 is what the YAML expects.
    for p in products:
        if p[0] == "COMP-3001":
            assert p[3] == 76, p


def build_us_labor_statistics_xlsx(out_path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    monthly = wb.active
    monthly.title = "Monthly Unemployment"
    # 12 months of 2024. July 2024 unemployment rate = 4.3%.
    # October 2024 has the lowest payroll add of 36,000.
    rows = [
        # Month        Unemp Rate (decimal),  Nonfarm payroll change
        ("January 2024",   0.037,  256_000),
        ("February 2024",  0.038,  269_000),
        ("March 2024",     0.038,  315_000),
        ("April 2024",     0.039,  175_000),
        ("May 2024",       0.040,  272_000),
        ("June 2024",      0.041,  206_000),
        ("July 2024",      0.043,  114_000),
        ("August 2024",    0.042,  142_000),
        ("September 2024", 0.041,  254_000),
        ("October 2024",   0.041,   36_000),
        ("November 2024",  0.042,  227_000),
        ("December 2024",  0.041,  256_000),
    ]
    monthly.append(["U.S. Bureau of Labor Statistics — 2024 Monthly Unemployment Summary"])
    monthly.append(["Unemployment Rate stored as decimal (e.g. 0.043 = 4.3%)"])
    monthly.append([])
    monthly.append(["Month", "Unemployment Rate", "Nonfarm Payroll Change"])
    for r in rows:
        monthly.append(list(r))
    monthly.append([])
    monthly.append(["Note: July 2024 unemployment rate is 4.3% (stored as 0.043)."])
    monthly.append([
        "Note: October 2024 had the lowest nonfarm payroll change of 2024 with 36,000 jobs added."
    ])
    monthly.append([
        "Note: This sheet contains aggregate monthly figures only — no breakdowns by age or other demographic groups."
    ])

    industry = wb.create_sheet("Industry Breakdown")
    industry.append(["U.S. Industry Employment — 2024 Year-over-Year Change"])
    industry.append([])
    industry.append(["Industry", "Jobs (Dec 2023)", "Jobs (Dec 2024)", "YoY Change"])
    industry_rows = [
        ("Manufacturing",                12_950_000, 12_850_000, -100_000),
        ("Health Care & Social Asst.",   21_700_000, 22_700_000, 1_000_000),
        ("Leisure & Hospitality",        16_800_000, 17_250_000,   450_000),
        ("Construction",                  8_100_000,  8_270_000,   170_000),
        ("Retail Trade",                 15_550_000, 15_490_000,   -60_000),
        ("Professional & Business Svcs.",22_950_000, 23_100_000,   150_000),
        ("Government",                   23_100_000, 23_500_000,   400_000),
        ("Mining & Logging",                640_000,    625_000,   -15_000),
        ("Information",                   3_100_000,  3_080_000,   -20_000),
        ("Financial Activities",          9_200_000,  9_245_000,    45_000),
    ]
    for r in industry_rows:
        industry.append(list(r))
    industry.append([])
    industry.append([
        "Note: Manufacturing posted the largest year-over-year job loss in 2024, with -100,000 jobs."
    ])
    industry.append([
        "Note: This sheet contains industry-level totals only — no demographic breakdowns by age or race."
    ])

    wb.save(out_path)


XLSX_BUILDERS = {
    "spreadsheets/company_financials_2024.xlsx": build_company_financials_xlsx,
    "spreadsheets/department_budget_2024.xlsx": build_department_budget_xlsx,
    "spreadsheets/product_inventory.xlsx": build_product_inventory_xlsx,
    "spreadsheets/us_labor_statistics_2024.xlsx": build_us_labor_statistics_xlsx,
}


# ============================================================================
# DRIVER
# ============================================================================
def main() -> None:
    print(f"Writing real-world corpus under {CORPUS_ROOT}")
    written = 0

    # Text docs
    for rel_path, body in TEXT_DOCS.items():
        out = CORPUS_ROOT / rel_path
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(body, encoding="utf-8")
        size = os.path.getsize(out)
        print(f"  [TXT ] {rel_path}  ({size:,} bytes)")
        written += 1

    # XLSX docs
    for rel_path, builder in XLSX_BUILDERS.items():
        out = CORPUS_ROOT / rel_path
        out.parent.mkdir(parents=True, exist_ok=True)
        builder(out)
        size = os.path.getsize(out)
        print(f"  [XLSX] {rel_path}  ({size:,} bytes)")
        written += 1

    print(f"=== Wrote {written} documents (expected 19) ===")
    assert written == 19, f"Expected 19 documents, wrote {written}"


if __name__ == "__main__":
    main()
